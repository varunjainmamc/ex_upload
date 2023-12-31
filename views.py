from distutils.command import sdist
from django.shortcuts import render, redirect
from django.contrib.auth.mixins import LoginRequiredMixin
from django.urls import reverse_lazy
import pandas as pd
import openpyxl
from io import StringIO
from django.views import View
from ex_upload.models import Table_E
from .forms import PasteDataForm, ExcelUploadForm, ConfirmationForm
from django.http import JsonResponse
from django.contrib import messages


def paste_data(request):
    if request.method == 'POST':
        form = PasteDataForm(request.POST)
        if form.is_valid():
            pasted_data = form.cleaned_data['data']
            # Process pasted_data, e.g., convert it to DataFrame
            df = pd.read_csv(StringIO(pasted_data), sep='\t')  # Adjust separator as needed
            # Do something with the DataFrame 'df', e.g., save to database
            # ...

            return render(request, 'ex_upload/paste_success.html', {'df': df})
    else:
        form = PasteDataForm()

    return render(request, 'ex_upload/paste_data.html', {'form': form})
    
class input_feedbackView(View):
    template_feedback = 'ex_upload/feedback.html'
    template_demo = 'ex_upload/demo.html'
    template_input = 'ex_upload/input.html'
    # success_url = reverse_lazy('table_revision:input_data')
    def common_code(self,user,table_name):
        rows = Table_E.objects.filter(h1=user, h3=table_name) # Fetch rows based on username and table_name
        unique_key = rows.filter(h2='data').values_list('h4', flat=True)
        unique_key3 = rows.filter(h2='data').values_list('h6', flat=True).distinct()
        header_row = rows.filter(h2='header').first() # Extract the header from the row where 'header' column is 'Y'
        header = [getattr(header_row, f'h{i}', '') for i in range(4, 51)]
        num_row = rows.count()-1
        num_row_list = [str(i) for i in range(1, num_row+1)]
        # Remove empty columns
        header = [col for col in header if col]
        num_col = len(header)    
        ctx = {'header': header, 'num_col': num_col, 'num_row': num_row,\
                'table_name': table_name, 'num_row_list':num_row_list,\
                      'unique_key': unique_key, 'unique_key3': unique_key3, 'rows':rows}
        return ctx
    
    def get(self, request):
        if not request.user.is_authenticated: # demo table
            user = 'varunjainmamc'        
            table_name = 'demo'
            ctx = self.common_code(user, table_name)
            feedback = [(f'h{i}',"") for i in range(4, ctx['num_col']+4)]
            # print(feedback) 
            ctx['feedback'] = feedback
            ctx['frm'] = 'frm1'
            #print("ctx1", ctx)
            return render(request, self.template_demo, ctx)
        else:
            # page to get existing tables as drop down for user input
            table_all = Table_E.objects.filter(h1=request.user.username).values_list('h3', flat=True).distinct()
            ctx = {'table_all': table_all}
            return render(request, self.template_input, ctx)

    def post(self, request):
        if not request.user.is_authenticated: # feedback on demo table input
            user = request.POST.get('user')
        else:        
            user = request.user.username  # feedback on user table input    

        table_name = request.POST['table_name'] # user input and feedback : common code    
        ctx = self.common_code(user, table_name) 
        if request.POST['page_t']: # post req to display empty col for user input
            num_col = ctx['num_col']
            feedback = [(f'h{i}',"") for i in range(4, num_col+4)]
            ctx['feedback'] = feedback
            return render(request, self.template_input, ctx)
        
        else: # post req to display feedback
            feedback ={}
            for j in range(1,ctx['num_row']+1):
                feedback[j] ={} 
                key = f'h4{j}'
                #print('k2',key)
                ukv = request.POST.get(key, '')
                #print('ukv',ukv)
                existing_data = ctx['rows'].filter(h4__iexact=ukv).first()
                #print(existing_data)
                row = {} # row is dic to collect user value of every current iterating row against field name
                if existing_data:                
                    for f_name, f_value in request.POST.items():
                        if f_name[2:] == str(j):
                            row[f_name] = f_value
                    for field_name, value_form in row.items():
                        modified_field_name = field_name[:-1]
                        correct_value = getattr(existing_data, modified_field_name)
                        #print('correct_value', correct_value)
                        if str(value_form).lower() == str(correct_value).lower():
                            feedback[j][field_name] = [value_form,1] # 1 will used in html to display "correct"
                        else:
                            feedback[j][field_name] = [value_form,correct_value]
            ctx['feedback'] = feedback
            ctx['frm'] = 'frm2'
            ctx['initial_anchor'] = 'targetSection'
            #print("ctx2", ctx)
            if not request.user.is_authenticated: # demo table
                return render(request, self.template_demo, ctx)
            else:
                return render(request, self.template_feedback, ctx)
            
class ConfirmationView(View):
    template_name = 'ex_upload/confirmation.html'
    # template_name = 'alt.html'
    def get(self, request): # display data uploaded from excel
        # Retrieve the dictionary from the session
        data_dict = request.session.get('data_dict', {})
        table_name = request.session['table_name']
        # print('session data')
        # print(data_dict)
        ctx= {'data_dict': data_dict, 'table_name' : table_name}
        return render(request, self.template_name, ctx)

    def post(self, request): # upon cofirmation
        table_name = request.POST['table_name']
        upload_dict = dict(request.POST.items())
        df_dict = {key: value for key, value in upload_dict.items() if 'fieup' in key }
        # print('df_dict')
        for key, value in df_dict.items():
          print(f'{key}: {value}')
        # Convert dictionary items to a list of tuples
        data_tuples = [(int(key.split("fieup")[0]), int(key.split("fieup")[1]), value) for key, value in df_dict.items()]
        df = pd.DataFrame(data_tuples, columns=['Row', 'Col', 'Value']) # Create a DataFrame from the list of tuples
        print (df)
        df = df.pivot(index='Row', columns='Col', values='Value') # Pivot the DataFrame to get the desired format
        df.reset_index(drop=True, inplace=True) # Reset the index if needed
        # print(df)
        
        # Reconstruct the DataFrame
#         df = pd.DataFrame(data_dict['data'], columns=data_dict['columns'])
        df.insert(0, 'User_Name', request.user.username)
        df.insert(1, 'header', "data")
        df.loc[0, 'header'] = "header"
        df.insert(2, 'table_name', table_name)            
        num_col = 53- len(df.columns)
        if num_col > 0:
            for i in range(num_col):
                df.insert(53-num_col+i,f'hi{i}','')
        df.columns = [f'h{i}' for i in range(1, 54)]
        # print('df.columns',df.columns)            
        data_to_insert = df.to_dict(orient='records') # Convert DataFrame to a list of dictionaries
        Table_E.objects.bulk_create([Table_E(**data) for data in data_to_insert]) # Bulk insert into the database
        request.session.pop('df', None) # Clear the session data
        return redirect('ex_upload:input') # Redirect to a success page

class upload_excel(LoginRequiredMixin,View):
    def get(self, request):  # display upload excel field      
        form = ExcelUploadForm()
        table_all = Table_E.objects.filter(h1=request.user.username).values_list('h3', flat=True).distinct()
        ctx = {'table_all': table_all, 'form': form}
        return render(request, 'ex_upload/upload_excel.html', ctx)

    def post(self, request):  # Process uploaded excel data
        form = ExcelUploadForm(request.POST, request.FILES)
        if form.is_valid():
            table_name = form.cleaned_data['table_name']
            request.session['table_name'] = table_name
            excel_file = request.FILES['excel_file']

            df = pd.read_excel(excel_file, header=None)  # Step 1
            num_rows = df.shape[0]
            num_cols = df.shape[1]
            num_rows = min(num_rows, 50)
            num_cols = min(num_cols, 50)
            # print ("df num_rows",num_rows,num_cols)
            df = df.iloc[:num_rows, :num_cols]
            book = openpyxl.load_workbook(excel_file) # Step 2: Extract cell formatting information
            sheet = book.active
            cell_form = {} # Step 3: Determine the number of decimal places visible in each Excel cell
            for col_index in range(1, num_cols+1):
                # print ("cell form for col", col_index)
                col_letter = openpyxl.utils.get_column_letter(col_index)
                for row_index, cell in enumerate(sheet.iter_rows(min_row=1, min_col=col_index, max_col=col_index,\
                                                                  max_row=num_rows), start=1):
                    # if col_index == 1:
                    #     print ("cell form for row", row_index)
                    cell_value = str(cell[0].value)
                    number_format = sheet.cell(row=row_index, column=col_index).number_format
                    cell_form[(row_index, col_index)]={}
                    cell_form[(row_index, col_index)]['format'] = number_format
                    try:
                        cell_form[(row_index, col_index)]['length'] = len(number_format.split('.')[-1])  if '.' in number_format else (1 if '%' in number_format else 0)
                    except AttributeError:
                        cell_form[(row_index, col_index)]['length'] = 0

                    # print("chec", row_index, col_index, cell_value, cell_form[(row_index, col_index)])
            for (row_index, col_index), places in cell_form.items(): # Step 4: Round the corresponding DataFrame columns based on the determined decimal places
                length = cell_form[(row_index, col_index)]['length']
                format = cell_form[(row_index, col_index)]['format']
                # print(row_index, col_index, length, format)
                cell_value = df.iloc[row_index - 1, col_index - 1]
                if '%' in format:
                    cell_value = f"{cell_value:.{length-1}%}"
                elif '0' in format:
                    rounded_value = round(cell_value, length)
                    cell_value = f"{rounded_value:.{length}f}" # store the numeric value based on the visible decimal places in excel
                df.iloc[row_index - 1, col_index - 1] = str(cell_value) # Update the DataFrame with the rounded value

            # Check if the table exists
            table_exists = Table_E.objects.filter(h1=request.user.username, h3=table_name).exists()
            if table_exists:
                # Get the header saved in the database
                db_header_row = Table_E.objects.filter(h1=request.user.username, h2='header', h3=table_name).first()
                db_header_columns = [getattr(db_header_row, f'h{i}') for i in range(4, 51) if\
                                      getattr(db_header_row, f'h{i}') is not None and\
                                          getattr(db_header_row, f'h{i}') != '']

                excel_header_columns = [df.iloc[0, i] for i in range(df.shape[1])]
                # Compare the headers
                if db_header_columns != excel_header_columns:
                    messages.warning(request, 'Warning: The header in the Excel file does not match the saved header in the database.')
                    return redirect('ex_upload:upload')  # Adjust the namespace and URL name according to your project
                  
                for index, row in df.iloc[1:].iterrows():
                    # Get the value in the first column (assuming it's the 'h4' column)
                    value_to_match = row[0]

                    # Check if the value exists in the database
                    existing_row = Table_E.objects.filter(h1=request.user.username, h3=table_name, h4=value_to_match).first()

                    if existing_row:
                        # If the row exists, update its values
                        for i in range(1, df.shape[1]):
                            setattr(existing_row, f'h{i+4}', row[i])
                        existing_row.save()
                    else:
                        # If the row doesn't exist, create a new row
                        new_row = Table_E(h1=request.user.username, h2='data', h3=table_name, **{f'h{i+4}': row[i] for i in range(df.shape[1])})
                        new_row.save()
                    messages.success(request, 'Data successfully processed.')
                    return redirect('ex_upload:input')
            else:
                data_dict = df.to_dict(orient='split')
                request.session['data_dict'] = data_dict
                return redirect('ex_upload:confirmation') 
        return render(request, 'ex_upload/upload_excel.html', {'form': form})   

class DeleteView(LoginRequiredMixin,View):
    template_name = 'ex_upload/delete.html'
    
    def get(self, request):
        form = ExcelUploadForm()
        form.fields['table_name'].widget.attrs['placeholder'] = 'Placeholder text for Page 1'
        table_all = Table_E.objects.filter(h1=request.user.username).values_list('h3', flat=True).distinct()
        ctx = {'table_all': table_all, 'form': form}        
        return render(request, self.template_name, ctx)
    
    def post(self, request):
        table_name = request.POST.get('table_name')
        Table_E.objects.filter(h1=request.user.username, h3=table_name).delete()
        return redirect('ex_upload:delete') # Redirect to a success page
